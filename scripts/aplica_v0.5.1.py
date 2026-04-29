#!/usr/bin/env python3
"""Patch idempotente do template Calculo.xlsx — Bloco A v0.5.1 (openpyxl).

Executa A1–A11 na ordem; imprime relatório A12 ao final.
"""

from __future__ import annotations

import shutil
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = PROJECT_ROOT / "templates" / "Calculo.xlsx"
BACKUP = PROJECT_ROOT / "templates" / "Calculo.original.xlsx"

LIMITES = {
    "PRICE 36x": 168,
    "PRICE 48x": 180,
    "PRICE 60x": 192,
}

PRINT_ROW_END = {
    "PRICE 24X": 155,
    "PRICE 36x": 168,
    "PRICE 48x": 180,
    "PRICE 60x": 192,
}


def copy_cell_style(src: Cell, dst: Cell) -> None:
    """Preserva font, fill, border, alignment, number_format, protection."""
    if src is None or dst is None:
        return
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def _backup_if_needed() -> None:
    if not BACKUP.exists():
        shutil.copy2(TEMPLATE, BACKUP)
        print(f"[backup] Criado {BACKUP.relative_to(PROJECT_ROOT)}")
    else:
        print(f"[backup] Já existe {BACKUP.relative_to(PROJECT_ROOT)} — não sobrescreve")


def aplicar_a1_a2(wb) -> None:
    # A1 — PRICE 24X
    sh = wb["PRICE 24X"]
    sh["C5"] = "Data do 1º vencimento:"
    sh["D5"].value = None
    sh["D5"].number_format = "dd/mm/yyyy"
    copy_cell_style(sh["C3"], sh["C5"])
    copy_cell_style(sh["D3"], sh["D5"])

    # A2 — outras abas
    for nome_aba in ["PRICE 36x", "PRICE 48x", "PRICE 60x"]:
        sh = wb[nome_aba]
        sh["C6"] = "Data do 1º vencimento:"
        sh["D6"].value = None
        sh["D6"].number_format = "dd/mm/yyyy"
        copy_cell_style(sh["C4"], sh["C6"])
        copy_cell_style(sh["D4"], sh["D6"])


def aplicar_a3(wb) -> None:
    sh = wb["PRICE 24X"]
    for new_row in range(147, 156):
        for col_idx in range(1, sh.max_column + 1):
            src = sh.cell(row=132, column=col_idx)
            dst = sh.cell(row=new_row, column=col_idx)
            copy_cell_style(src, dst)
        # Coluna C — número da parcela
        sh.cell(row=new_row, column=3).value = new_row - 131
        # AE / BF replicam D (colunas 31 e 58 para AE/BF — usar nomes)
        sh[f"AE{new_row}"] = f"=D{new_row}"
        sh[f"BF{new_row}"] = f"=AE{new_row}"


def aplicar_a4(wb) -> None:
    sh = wb["PRICE 24X"]
    sh["D132"] = '=IF(C132<=$I$15, $D$5, "")'
    for row in range(133, 156):
        sh[f"D{row}"] = f'=IF(C{row}<=$I$15, EDATE($D$5, C{row}-1), "")'


def aplicar_a5(wb) -> None:
    for nome_aba, fim in LIMITES.items():
        sh = wb[nome_aba]
        sh["D133"] = '=IF(C133<=$I$16, $D$6, "")'
        for row in range(134, fim + 1):
            sh[f"D{row}"] = f'=IF(C{row}<=$I$16, EDATE($D$6, C{row}-1), "")'


def aplicar_a6(wb) -> None:
    sh = wb["PRICE 24X"]
    for row in range(132, 156):
        sh[f"N{row}"] = f'=IF(C{row}<=$I$15, $I$16, "")'

    for nome_aba, fim in LIMITES.items():
        sh = wb[nome_aba]
        for row in range(133, fim + 1):
            sh[f"N{row}"] = f'=IF(C{row}<=$I$16, $I$17, "")'


def aplicar_a7(wb) -> None:
    celulas_data = {
        "PRICE 24X": ["D3", "D4", "D5"]
        + [f"D{r}" for r in range(132, 156)]
        + [f"AE{r}" for r in range(132, 156)]
        + [f"BF{r}" for r in range(132, 156)],
        "PRICE 36x": ["D4", "D5", "D6"]
        + [f"D{r}" for r in range(133, 169)]
        + [f"AE{r}" for r in range(133, 169)]
        + [f"BF{r}" for r in range(133, 169)],
        "PRICE 48x": ["D4", "D5", "D6"]
        + [f"D{r}" for r in range(133, 181)]
        + [f"AE{r}" for r in range(133, 181)]
        + [f"BF{r}" for r in range(133, 181)],
        "PRICE 60x": ["D4", "D5", "D6"]
        + [f"D{r}" for r in range(133, 193)]
        + [f"AE{r}" for r in range(133, 193)]
        + [f"BF{r}" for r in range(133, 193)],
    }
    us_like_substrings = ("mm-dd-yy", "mm/dd", "m/d/yy", "m/d/y")

    for nome_aba, celulas in celulas_data.items():
        sh = wb[nome_aba]
        for ref in celulas:
            sh[ref].number_format = "dd/mm/yyyy"

        # Varredura defensiva: troca formatos US óbvios em toda a aba
        for row in sh.iter_rows():
            for cell in row:
                nf = (cell.number_format or "").lower()
                if any(s in nf for s in us_like_substrings):
                    cell.number_format = "dd/mm/yyyy"


def aplicar_a8(wb) -> None:
    sh = wb["PRICE 48x"]
    sh.page_setup.orientation = sh.ORIENTATION_LANDSCAPE
    sh.page_setup.paperSize = sh.PAPERSIZE_A4


def aplicar_a9(wb) -> None:
    wb["PRICE 60x"]["D4"].value = None


def aplicar_a11(wb) -> None:
    try:
        from openpyxl.worksheet.properties import PageSetupProperties
    except ImportError:
        PageSetupProperties = None  # type: ignore

    for nome_aba in ["PRICE 24X", "PRICE 36x", "PRICE 48x", "PRICE 60x"]:
        sh = wb[nome_aba]
        sh.page_setup.orientation = sh.ORIENTATION_LANDSCAPE
        sh.page_setup.paperSize = sh.PAPERSIZE_A4
        sh.page_setup.fitToWidth = 1
        sh.page_setup.fitToHeight = 0
        if PageSetupProperties is not None:
            if sh.sheet_properties.pageSetUpPr is None:
                sh.sheet_properties.pageSetUpPr = PageSetupProperties()
            sh.sheet_properties.pageSetUpPr.fitToPage = True
        sh.page_margins.left = 0.7 / 2.54
        sh.page_margins.right = 0.7 / 2.54
        sh.page_margins.top = 0.7 / 2.54
        sh.page_margins.bottom = 0.7 / 2.54
        sh.page_margins.header = 0.3 / 2.54
        sh.page_margins.footer = 0.3 / 2.54

        end = PRINT_ROW_END[nome_aba]
        sh.print_area = f"A1:CC{end}"

        sh.oddHeader.center.text = '&"Calibri,Bold"&14Cálculo de Ação Crefaz'
        sh.oddFooter.right.text = "Página &P de &N"


def validar_a12(wb) -> list[str]:
    """Retorna lista de erros (vazia = OK)."""
    erros: list[str] = []
    sh = wb["PRICE 24X"]
    if sh["D5"].value is not None:
        erros.append(f"D5 deveria ser None, veio {sh['D5'].value!r}")
    if (sh["D5"].number_format or "") != "dd/mm/yyyy":
        erros.append(f"D5 number_format esperado dd/mm/yyyy, veio {sh['D5'].number_format!r}")

    v132 = sh["D132"].value
    if not isinstance(v132, str) or not v132.startswith("=IF(C132"):
        erros.append(f"D132 fórmula IF esperada, veio {v132!r}")

    if sh["C155"].value != 24:
        erros.append(f"C155 esperado 24, veio {sh['C155'].value!r}")

    orient = sh.page_setup.orientation
    if orient not in ("landscape", sh.ORIENTATION_LANDSCAPE):
        erros.append(f"Orientação PRICE 24X: esperado landscape, veio {orient!r}")

    ftw = sh.page_setup.fitToWidth
    if ftw not in (1, True):
        erros.append(f"fitToWidth esperado 1, veio {ftw!r}")

    return erros


def main() -> int:
    if not TEMPLATE.exists():
        print(f"ERRO: template não encontrado: {TEMPLATE}", file=sys.stderr)
        return 2

    print(f"[load] {TEMPLATE}")
    wb = load_workbook(TEMPLATE)

    _backup_if_needed()

    aplicar_a1_a2(wb)
    aplicar_a3(wb)
    aplicar_a4(wb)
    aplicar_a5(wb)
    aplicar_a6(wb)
    aplicar_a7(wb)
    aplicar_a8(wb)
    aplicar_a9(wb)
    aplicar_a11(wb)

    wb.save(TEMPLATE)
    print(f"[save] {TEMPLATE.relative_to(PROJECT_ROOT)}")

    wb2 = load_workbook(TEMPLATE)
    erros = validar_a12(wb2)

    sh = wb2["PRICE 24X"]
    print("--- A12 ---")
    print("D5:", sh["D5"].value, "|", sh["D5"].number_format)
    print("D132:", sh["D132"].value, "|", sh["D132"].number_format)
    print("D155:", sh["D155"].value, "|", sh["D155"].number_format)
    print("C155:", sh["C155"].value)
    print("N132:", sh["N132"].value)
    print("Page orientation:", sh.page_setup.orientation)
    print("Fit to width:", sh.page_setup.fitToWidth)

    if erros:
        print("\n[A12 FALHOU]")
        for e in erros:
            print(" -", e)
        return 1

    print("\n[A12 OK]")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
