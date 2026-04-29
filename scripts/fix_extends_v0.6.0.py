"""Patch idempotente — corrige rows 147-155 estendidas em v0.5.1 que ficaram
incompletas (faltavam merges de cells e fórmulas em colunas AO e BP).

Sintoma corrigido: ao gerar PDF via LibreOffice headless, parcelas 16-24 nas
tabelas RECALCULADAS e VALORES PAGOS apareciam visualmente vazias mesmo com
cache de fórmulas correto. Causa: o `aplica_v0.5.1.py` original copiou
borders/fonts mas não replicou (a) merges de blocos D-M, N-Y, AE-AN, AO-AZ,
BF-BO, BP-CC; nem (b) fórmulas `=AO_anterior` e `=BP_anterior`.

Rodar em qualquer Calculo.xlsx que tenha aba CÁLCULO ou PRICE 24X.
Idempotente: ignora merges/fórmulas que já existem.

Uso:
    python3 scripts/fix_extends_v0.6.0.py [path/to/Calculo.xlsx]
    (default: templates/Calculo.xlsx)
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

GRUPOS_COL_MERGED = [
    ("D", "M"),
    ("N", "Y"),
    ("AE", "AN"),
    ("AO", "AZ"),
    ("BF", "BO"),
    ("BP", "CC"),
]

ROWS_ESTENDIDAS = range(147, 156)  # parcelas 16-24


def aplicar_fix(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path)
    aba = "CÁLCULO" if "CÁLCULO" in wb.sheetnames else "PRICE 24X"
    if aba not in wb.sheetnames:
        raise SystemExit(
            f"Aba 'CÁLCULO' nem 'PRICE 24X' encontrada. Abas: {wb.sheetnames}"
        )
    ws = wb[aba]

    relatorio = {
        "aba": aba,
        "merges_adicionadas": [],
        "formulas_AO": [],
        "formulas_BP": [],
    }

    # 1. Merges em rows 147-155
    for row in ROWS_ESTENDIDAS:
        for col_ini, col_fim in GRUPOS_COL_MERGED:
            rng = f"{col_ini}{row}:{col_fim}{row}"
            ja_existe = any(str(mr) == rng for mr in ws.merged_cells.ranges)
            if not ja_existe:
                ws.merge_cells(rng)
                relatorio["merges_adicionadas"].append(rng)

    # 2. Fórmulas AO/BP — cada row referencia a anterior (sequência)
    for row in ROWS_ESTENDIDAS:
        if ws[f"AO{row}"].value is None:
            ws[f"AO{row}"] = f"=AO{row-1}"
            relatorio["formulas_AO"].append(f"AO{row}")
        if ws[f"BP{row}"].value is None:
            ws[f"BP{row}"] = f"=BP{row-1}"
            relatorio["formulas_BP"].append(f"BP{row}")

    wb.save(xlsx_path)
    return relatorio


def main():
    default = Path("templates/Calculo.xlsx")
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default
    if not xlsx_path.exists():
        raise SystemExit(f"Arquivo não encontrado: {xlsx_path}")

    rel = aplicar_fix(xlsx_path)
    print(f"=== Patch v0.6.0 aplicado em {xlsx_path} ===")
    print(f"Aba: {rel['aba']}")
    print(f"Merges adicionadas: {len(rel['merges_adicionadas'])}")
    for m in rel["merges_adicionadas"]:
        print(f"  + {m}")
    print(f"Fórmulas AO adicionadas: {len(rel['formulas_AO'])}")
    print(f"Fórmulas BP adicionadas: {len(rel['formulas_BP'])}")


if __name__ == "__main__":
    main()
