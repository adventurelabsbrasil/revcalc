"""Gera templates/Calculo.quitado.xlsx a partir do CALCULO QUITADO.xlsx (hand-filled).

O arquivo de origem é uma planilha preenchida à mão (sample VANDRIANA na 24X), com:
  - datas de vencimento HARDCODED (não fórmula) e tabela truncada no nº de parcelas do sample;
  - células de input com valores de amostra;
  - abas 36x/48x/60x sem dados de amostra mas com o cálculo presente.

Este script a transforma num TEMPLATE input-driven, espelhando o padrão JÁ PROVADO da
aba ativa CÁLCULO (fórmulas guardadas por `IF(idx>$I$qtd,"")` + datas via EDATE ancoradas
no 1º vencimento), em TODAS as 4 abas e até a capacidade total de cada uma (24/36/48/60).

PRESERVA intacta a matemática de recálculo (bloco PMT, AH<n>, colunas AP/BL/saldo) —
só reescreve a LISTA de parcelas (3 tabelas lado a lado) e limpa os inputs de amostra.

Uso:
    python scripts/limpa_template_quitado.py "/caminho/CALCULO QUITADO.xlsx"

Validação numérica final é E2E no xeon (LibreOffice) contra a base EDUARDO CESAR (1956714).
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from calculadora_crefaz.config import celulas_para_aba

_RE_COORD = re.compile(r"^([A-Z]+)(\d+)$")


def _abs(coord: str) -> str:
    """'I15' → '$I$15'; 'BL8' → '$BL$8' (lida com colunas de 2 letras)."""
    m = _RE_COORD.match(coord)
    if not m:
        raise ValueError(f"Coordenada inválida: {coord!r}")
    return f"${m.group(1)}${m.group(2)}"

# Parâmetros por aba (derivados por inspeção — ver docs/sessions).
#   off       : offset de linha vs 24X (layout +1 nas longas)
#   data0     : 1ª linha de dados da lista de parcelas
#   cap       : capacidade de parcelas (= prazo máximo daquela aba)
#   qtd/pp/pv : células de qtd parcelas / parcelas pagas / 1º vencimento (âncoras)
#   prest     : célula do valor da prestação ($I$..)
#   ao_first  : célula do valor da parcela recalculada (1ª ref da coluna AO)
# res_row = linha vazia do bloco-resumo onde injetamos "Valor da parcela
#   recalculada = " (p/ a imag.02 resumo incluir o resultado sem a derivação PMT).
# res_src = célula que já guarda a parcela recalculada (=AH<n> via AP16/AP17).
# fin_row = linha "Valor financiado ajustado" (fonte do estilo/formato do valor).
ABAS = {
    "PRICE 24X": dict(off=0, data0=132, cap=24, qtd="I15", pp="BL8", pv="D5", prest="I16", ao_first="AH121", res_row=32, res_src="AP16", fin_row=29),
    "PRICE 36x": dict(off=1, data0=133, cap=36, qtd="I16", pp="BL9", pv="D6", prest="I17", ao_first="AH122", res_row=33, res_src="AP17", fin_row=30),
    "PRICE 48x": dict(off=1, data0=133, cap=48, qtd="I16", pp="BL9", pv="D6", prest="I17", ao_first="AH122", res_row=33, res_src="AP17", fin_row=30),
    "PRICE 60x": dict(off=1, data0=133, cap=60, qtd="I16", pp="BL9", pv="D6", prest="I17", ao_first="AH122", res_row=33, res_src="AP17", fin_row=30),
}


def _injeta_resultado_resumo(ws, p: dict) -> None:
    """Adiciona 'Valor da parcela recalculada = <=AP16/17>' no bloco-resumo da imag.02.

    Espelha a linha "Taxa média" (label simples + valor mergeado AI:AR), com o
    formato contábil do "Valor financiado ajustado" (mostra 103,07 sem ##).
    """
    from copy import copy
    rr, fin = p["res_row"], p["fin_row"]
    taxa = fin + 1  # linha "Taxa média de juros ao mês" — label single-line + valor merged
    # Label (single-line, transborda à direita como o da taxa)
    lbl = ws[f"AD{rr}"]
    lbl.value = "Valor da parcela recalculada = "
    src_lbl = ws[f"AD{taxa}"]
    lbl.font = copy(src_lbl.font)
    lbl.alignment = copy(src_lbl.alignment)
    # Valor: mesmo merge da linha da taxa (AI:AR) + formato contábil do financiado
    ws.merge_cells(f"AI{rr}:AR{rr}")
    val = ws[f"AI{rr}"]
    val.value = f"={p['res_src']}"
    src_val = ws[f"AI{fin}"]
    val.font = copy(src_val.font)
    val.alignment = copy(src_val.alignment)
    val.number_format = src_val.number_format


def _reescreve_lista_parcelas(ws, p: dict) -> None:
    data0, cap = p["data0"], p["cap"]
    qtd, pp, pv, prest = _abs(p["qtd"]), _abs(p["pp"]), _abs(p["pv"]), _abs(p["prest"])
    ao_first = p["ao_first"]
    idx_base = data0 - 1          # ROW()-idx_base == índice da parcela (1..)
    edate_base = data0            # ROW()-edate_base == k meses (0 na 1ª parcela)
    idx = f"ROW()-{idx_base}"
    g = f'IF({idx}>{qtd},"",'     # prefixo do guard; cada fórmula fecha com ')'
    last = data0 + cap - 1
    for r in range(data0, last + 1):
        # PARCELAS CONFORME O CONTRATO
        ws[f"C{r}"] = f"={g}{idx})"
        ws[f"D{r}"] = f"={g}EDATE({pv},ROW()-{edate_base}))"
        ws[f"N{r}"] = f"={g}{prest})"
        # PARCELAS RECALCULADAS
        ws[f"AD{r}"] = f"={g}{idx})"
        ws[f"AE{r}"] = f"=D{r}"
        ao_ref = ao_first if r == data0 else f"AO{r-1}"
        ws[f"AO{r}"] = f"={g}{ao_ref})"
        # VALORES PAGOS PELO CLIENTE
        ws[f"BE{r}"] = f'={g}IF({idx}<={pp},{idx},""))'
        ws[f"BF{r}"] = f'=IF(BE{r}="","",AE{r})'
        ws[f"BP{r}"] = f'={g}IF({idx}<={pp},{prest},""))'


def _limpa_inputs(ws, aba: str) -> None:
    """Zera as células de input (serão preenchidas em runtime por _preencher_aba)."""
    for coord in celulas_para_aba(aba).values():
        ws[coord] = None


def main(origem: str) -> None:
    raiz = Path(__file__).resolve().parents[1]
    destino = raiz / "templates" / "Calculo.quitado.xlsx"
    wb = load_workbook(origem)
    if list(wb.sheetnames) != list(ABAS.keys()):
        print(f"AVISO: abas inesperadas {wb.sheetnames} (esperado {list(ABAS)})")
    for aba, p in ABAS.items():
        ws = wb[aba]
        _limpa_inputs(ws, aba)
        _reescreve_lista_parcelas(ws, p)
        _injeta_resultado_resumo(ws, p)
        print(f"  [{aba}] lista de parcelas {p['data0']}..{p['data0']+p['cap']-1} guardada; inputs limpos; resultado no resumo (AD{p['res_row']})")
    wb.save(destino)
    print(f"OK → {destino}")


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else str(Path.home() / "Downloads" / "CALCULO QUITADO.xlsx")
    main(src)
