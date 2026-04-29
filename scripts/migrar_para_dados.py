"""Migra `templates/Calculo.xlsx` para arquitetura DADOS+visual.

Pré-condição:
- v0.5.1 cirúrgica já aplicada (tabela 24X estendida, IF+EDATE, page setup uniforme).
- Aba `DADOS` ainda não existe.

Ações:
1. Insere aba `DADOS` na primeira posição com:
   - Coluna A: rótulos pt-BR (A2..A14).
   - Coluna B: vazia, com number_format por tipo.
2. Para cada PRICE (24X, 36x, 48x, 60x), substitui o VALOR das células de input
   (mapa de `celulas_para_aba` em config.py) por fórmula `=DADOS!$B$X`.
3. Esconde a aba DADOS (sheet_state='hidden') — Ctrl+P só imprime PRICE.
4. NÃO toca nas fórmulas da tabela de parcelas (D132+/N132+/AE/BF) — elas
   referenciam $D$5/$D$6/$I$15/$I$16/$I$17 e Excel resolve transitivamente
   após D5/D6/I... virarem fórmulas DADOS.

Idempotente: pode rodar 2x sem corromper. Se aba DADOS já existe, aborta.

Uso:
    python3 scripts/migrar_para_dados.py <input.xlsx> <output.xlsx>
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Mapa DADOS — espelha CELULAS_DADOS de calculadora_crefaz.config
DADOS_LABELS_E_FORMATOS = [
    ("B2",  "titulo",              "Título da operação",         "@",                                  "text"),
    ("B3",  "data_emissao",        "Data de emissão",            "dd/mm/yyyy",                         "date"),
    ("B4",  "primeiro_vencimento", "1º vencimento",              "dd/mm/yyyy",                         "date"),
    ("B5",  "ultimo_vencimento",   "Último vencimento",          "dd/mm/yyyy",                         "date"),
    ("B6",  "valor_principal",     "Valor principal",            'R$ #,##0.00',                        "currency"),
    ("B7",  "tac",                 "TAC / tarifas",              'R$ #,##0.00',                        "currency"),
    ("B8",  "iof",                 "IOF",                        'R$ #,##0.00',                        "currency"),
    ("B9",  "qtd_parcelas",        "Quantidade de parcelas",     "0",                                  "int"),
    ("B10", "valor_prestacao",     "Valor da prestação",         'R$ #,##0.00',                        "currency"),
    ("B11", "taxa_pactuada",       "Taxa pactuada (mensal)",     "0.00%",                              "percent"),
    ("B12", "taxa_bacen",          "Taxa BACEN (mensal)",        "0.00%",                              "percent"),
    ("B13", "parcelas_pagas",      "Parcelas pagas",             "0",                                  "int"),
    ("B14", "data_calculo",        "Data do cálculo",            "dd/mm/yyyy",                         "date"),
]

# Mapeamento campo -> célula DADOS (espelha CELULAS_DADOS)
CAMPO_PARA_DADOS = {item[1]: item[0] for item in DADOS_LABELS_E_FORMATOS}

# Mapeamento campo -> célula DADOS pra fórmula =DADOS!$B$X
def formula_dados(campo: str) -> str:
    ref = CAMPO_PARA_DADOS[campo]
    col = "".join(c for c in ref if c.isalpha())
    row = "".join(c for c in ref if c.isdigit())
    return f"=DADOS!${col}${row}"


# Espelha celulas_para_aba do config.py (offset +1 das outras abas vs PRICE 24X)
CELULAS_PRICE_24X = {
    "titulo": "C1",
    "data_pactuacao": "D3",          # mapeia pra data_emissao em DADOS
    "primeiro_vencimento": "D5",
    "valor_principal": "I7",
    "tac": "I9",
    "iof": "I13",
    "qtd_parcelas": "I15",
    "valor_prestacao": "I16",
    "taxa_pactuada": "I17",
    "taxa_bacen": "AP15",
    "parcelas_pagas": "BL8",
}

def celulas_para_aba(aba: str) -> dict[str, str]:
    if aba == "PRICE 24X":
        return dict(CELULAS_PRICE_24X)
    out = {}
    for k, ref in CELULAS_PRICE_24X.items():
        col = "".join(c for c in ref if c.isalpha())
        row = int("".join(c for c in ref if c.isdigit()))
        out[k] = f"{col}{row + 1}"
    return out


# Mapeamento campo PRICE -> campo DADOS
# Nota: data_pactuacao (PRICE) <- data_emissao (DADOS) — MESMO valor, nome diferente
CAMPO_PRICE_PARA_DADOS = {
    "titulo": "titulo",
    "data_pactuacao": "data_emissao",
    "primeiro_vencimento": "primeiro_vencimento",
    "valor_principal": "valor_principal",
    "tac": "tac",
    "iof": "iof",
    "qtd_parcelas": "qtd_parcelas",
    "valor_prestacao": "valor_prestacao",
    "taxa_pactuada": "taxa_pactuada",
    "taxa_bacen": "taxa_bacen",
    "parcelas_pagas": "parcelas_pagas",
}


def migrar(in_path: Path, out_path: Path) -> dict:
    wb = load_workbook(in_path)
    relatorio = {"abas_inicial": list(wb.sheetnames), "substituicoes": [], "warnings": []}

    if "DADOS" in wb.sheetnames:
        raise SystemExit("Aba DADOS já existe — abortando (script é idempotente, este caso significa que a migração já rodou).")

    abas_price_esperadas = {"PRICE 24X", "PRICE 36x", "PRICE 48x", "PRICE 60x"}
    if not abas_price_esperadas.issubset(wb.sheetnames):
        raise SystemExit(f"Abas PRICE faltando: {abas_price_esperadas - set(wb.sheetnames)}")

    # 1. Criar aba DADOS na primeira posição
    ws_dados = wb.create_sheet("DADOS", 0)

    # Cabeçalho
    ws_dados["A1"] = "Campo"
    ws_dados["B1"] = "Valor"
    for cell_ref in ("A1", "B1"):
        c = ws_dados[cell_ref]
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", start_color="DDDDDD")
        c.alignment = Alignment(horizontal="center")

    # Linhas A2:B14
    for ref_b, campo, label, fmt, _tipo in DADOS_LABELS_E_FORMATOS:
        ref_a = ref_b.replace("B", "A")
        ws_dados[ref_a] = label
        ws_dados[ref_a].font = Font(bold=False)
        ws_dados[ref_a].alignment = Alignment(horizontal="left")
        # B fica vazia mas com number_format aplicado
        ws_dados[ref_b].number_format = fmt
    # Larguras razoáveis pra leitura humana
    ws_dados.column_dimensions["A"].width = 32
    ws_dados.column_dimensions["B"].width = 22

    # 2. Substituir valores nas células de input das 4 PRICE por fórmulas =DADOS!...
    for aba in ("PRICE 24X", "PRICE 36x", "PRICE 48x", "PRICE 60x"):
        ws = wb[aba]
        celulas = celulas_para_aba(aba)
        for campo_price, ref_price in celulas.items():
            campo_dados = CAMPO_PRICE_PARA_DADOS[campo_price]
            valor_anterior = ws[ref_price].value
            formato_anterior = ws[ref_price].number_format
            ws[ref_price] = formula_dados(campo_dados)
            # Preserva number_format atual da PRICE — não sobrescrever
            relatorio["substituicoes"].append({
                "aba": aba,
                "campo": campo_price,
                "ref": ref_price,
                "antes": str(valor_anterior)[:60] if valor_anterior is not None else None,
                "depois": formula_dados(campo_dados),
                "fmt_preservado": formato_anterior,
            })

    # 3. Esconder aba DADOS
    ws_dados.sheet_state = "hidden"

    # 4. Salvar
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    relatorio["abas_final"] = list(wb.sheetnames)
    return relatorio


def main():
    if len(sys.argv) < 3:
        print("Uso: migrar_para_dados.py <input.xlsx> <output.xlsx>")
        sys.exit(2)
    in_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2])
    if not in_path.exists():
        raise SystemExit(f"Input não encontrado: {in_path}")

    rel = migrar(in_path, out_path)

    # Relatório
    print(f"\n=== Migração DADOS+visual concluída ===")
    print(f"Input : {in_path}")
    print(f"Output: {out_path}")
    print(f"Abas iniciais: {rel['abas_inicial']}")
    print(f"Abas finais  : {rel['abas_final']}")
    print(f"\nSubstituições feitas: {len(rel['substituicoes'])}")
    for s in rel["substituicoes"]:
        antes = s["antes"]
        if antes is None:
            antes = "(vazio)"
        print(f"  [{s['aba']:11s}] {s['ref']:6s} ({s['campo']:22s}) {antes!r:55s} → {s['depois']}")
    if rel["warnings"]:
        print(f"\nWarnings:")
        for w in rel["warnings"]:
            print(f"  ! {w}")
    else:
        print("\nWarnings: nenhum.")


if __name__ == "__main__":
    main()
