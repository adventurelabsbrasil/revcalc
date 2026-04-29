"""Atualiza linhas das 3 tabelas de parcelas no template Calculo.xlsx (aba CÁLCULO).

- PARCELAS CONFORME (C, D, N): parcela, vencimento e parcela cobrada só até `$I$15`
  linhas utilizáveis (prazo do contrato). Linhas extras ficam visualmente em branco.
- AE continua igual a D; cadeias AO/BF seguem intactas herdando apenas D válido onde houver linha útil.
- VALORES PAGOS (BP): só preenche `I$16` nas linhas onde o índice da parcela
  está entre 1 e o mínimo lógico: `indice <= $I$15` e `indice <= $BL$8`;

Idempotente: rodar duas vezes é seguro.

Uso:
    python scripts/patch_tabelas_parcelas_dynamic.py [templates/Calculo.xlsx]
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


ROW_FIRST = 132  # primeira linha de dados; parcela número 1
LAST = 155
FIRST_IDX = ROW_FIRST - 1  # 131 → ROW()-FIRST_IDX = parcela atual


def aplicar(planilha: Path) -> None:
    wb = load_workbook(planilha)
    if "CÁLCULO" not in wb.sheetnames:
        raise SystemExit(f"Aba 'CÁLCULO' ausente em {planilha}")
    ws = wb["CÁLCULO"]
    idx = f"ROW()-{FIRST_IDX}"
    ed_off = f"ROW()-{ROW_FIRST}"

    for r in range(ROW_FIRST, LAST + 1):
        # Coluna Parcela (#)
        ws[f"C{r}"] = f'=IF({idx}>$I$15,"",{idx})'
        # Vencimento (conforme contrato); EDATE sempre a partir de D5
        ws[f"D{r}"] = f'=IF({idx}>$I$15,"",EDATE($D$5,{ed_off}))'
        # Parcela cobrada (valor pactuado) — apenas linhas válidas até o prazo
        ws[f"N{r}"] = f'=IF({idx}>$I$15,"",$I$16)'
        # Valor efetivamente pago pelo cliente: só até parcelas_pagas ($BL$8)
        ws[f"BP{r}"] = (
            f'=IF({idx}>$I$15,"",IF({idx}<=$BL$8,$I$16,""))'
        )

    # PARCELAS RECALCULADAS — coluna AO: mesma lógica do prazo (nada além de I15)
    ws["AO132"] = f'=IF({idx}>$I$15,"",AH121)'
    for r in range(ROW_FIRST + 1, LAST + 1):
        idx_r = f"ROW()-{FIRST_IDX}"
        prev = r - 1
        ws[f"AO{r}"] = f'=IF({idx_r}>$I$15,"",AO{prev})'

    wb.save(planilha)


def main() -> None:
    default = Path("templates/Calculo.xlsx")
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else default
    if not path.exists():
        raise SystemExit(f"Arquivo não encontrado: {path}")
    aplicar(path)
    print(f"Ok: {path}")


if __name__ == "__main__":
    main()
