"""Gera capturas (PDF + PNG) do XLSX usando LibreOffice headless + pypdfium2.

Fluxo:
1. Recebe `xlsx_bytes`, salva em arquivo temporário.
2. Chama `soffice --headless --convert-to pdf` para gerar PDF (uma página por
   "página de impressão" da planilha). Aproveita `page_setup` do template
   (paisagem A4, fitToWidth=1, fitToHeight=1 → 1 página/aba esperada).
3. Rasteriza cada página do PDF como PNG via pypdfium2 (DPI configurável).
4. Devolve PDF + lista de PNGs em memória, prontos pra upload no Drive.

Não exige Excel instalado. Funciona em macOS/Linux/Windows com LibreOffice
disponível no PATH (`soffice` ou `libreoffice`).

Uso:
    from .capturas import gerar_capturas, CapturasGeradas

    cap = gerar_capturas(xlsx_bytes, dpi=150)
    cap.pdf_bytes        # PDF do XLSX inteiro (1 página/aba típico)
    cap.pngs             # list[CapturaPng] com (nome_arquivo, bytes, pagina)
"""
from __future__ import annotations

import logging
import os
import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)

# Locale BR forçado no LibreOffice headless — garante datas dd/mm/yyyy e moeda
# R$ 1.234,56 mesmo se o ambiente (container/host) subir em locale C/US. Exige o
# locale pt_BR.UTF-8 gerado no SO (ver server/Dockerfile).
_SOFFICE_ENV = {**os.environ, "LANG": "pt_BR.UTF-8", "LC_ALL": "pt_BR.UTF-8", "LANGUAGE": "pt_BR:pt"}


# ─── Modelos ───────────────────────────────────────────────────────────────


@dataclass
class CapturaPng:
    """Uma captura PNG (uma página rasterizada do PDF)."""

    nome_arquivo: str  # ex: "13 Print PRICE 24X p1.png" — o pipeline define
    bytes_: bytes
    pagina: int  # 1-indexed
    largura_px: int
    altura_px: int


@dataclass
class CapturasGeradas:
    """Pacote completo de capturas pra um XLSX."""

    pdf_bytes: bytes
    pngs: list[CapturaPng] = field(default_factory=list)

    @property
    def num_paginas(self) -> int:
        return len(self.pngs)


# ─── Erros ──────────────────────────────────────────────────────────────────


class CapturasError(Exception):
    """Falha em qualquer etapa da geração de capturas."""


class LibreOfficeIndisponivel(CapturasError):
    """`soffice`/`libreoffice` não está no PATH."""


# ─── Helpers ────────────────────────────────────────────────────────────────


def _encontrar_soffice() -> str:
    """Resolve o binário do LibreOffice headless.

    Procura `soffice` (Linux/macOS) e `libreoffice` (alguns Linux).
    """
    for nome in ("soffice", "libreoffice"):
        caminho = shutil.which(nome)
        if caminho:
            return caminho
    raise LibreOfficeIndisponivel(
        "LibreOffice não encontrado no PATH. "
        "Instale: `brew install --cask libreoffice` (macOS), "
        "`apt install libreoffice` (Linux), ou baixe do site oficial (Windows)."
    )


def _xlsx_para_pdf(xlsx_bytes: bytes, *, timeout_s: int = 60) -> bytes:
    """Converte XLSX → PDF via LibreOffice headless. Retorna os bytes do PDF."""
    soffice = _encontrar_soffice()

    with tempfile.TemporaryDirectory(prefix="cap_crefaz_") as tmp:
        tmp_dir = Path(tmp)
        xlsx_path = tmp_dir / "in.xlsx"
        xlsx_path.write_bytes(xlsx_bytes)

        cmd = [
            soffice,
            "--headless",
            "--convert-to", "pdf",
            "--outdir", str(tmp_dir),
            str(xlsx_path),
        ]
        try:
            res = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=timeout_s,
                check=True,
                env=_SOFFICE_ENV,
            )
        except subprocess.TimeoutExpired:
            raise CapturasError(
                f"LibreOffice excedeu {timeout_s}s convertendo XLSX → PDF."
            )
        except subprocess.CalledProcessError as e:
            raise CapturasError(
                f"LibreOffice falhou (exit {e.returncode}): {e.stderr or e.stdout}"
            )

        logger.debug("soffice stdout: %s", res.stdout)
        pdf_path = tmp_dir / "in.pdf"
        if not pdf_path.exists():
            raise CapturasError(
                f"PDF esperado em {pdf_path} não foi gerado. soffice stdout: {res.stdout}"
            )
        return pdf_path.read_bytes()


def _pdf_para_pngs(pdf_bytes: bytes, *, dpi: int = 150) -> list[tuple[bytes, int, int]]:
    """Rasteriza PDF → lista de (png_bytes, largura_px, altura_px) — 1 entrada por página.

    Usa pypdfium2 (binding lightweight de PDFium, sem dependência de Poppler).
    """
    try:
        import pypdfium2 as pdfium
    except ImportError as e:
        raise CapturasError(
            f"pypdfium2 não instalado: {e}. Adicione ao requirements.txt."
        )

    import io
    from PIL import Image  # noqa: F401  # validar disponibilidade

    pdf = pdfium.PdfDocument(pdf_bytes)
    saidas: list[tuple[bytes, int, int]] = []
    scale = dpi / 72.0  # 72 DPI = referência interna do PDF
    for pagina in pdf:
        img = pagina.render(scale=scale).to_pil()
        buf = io.BytesIO()
        img.save(buf, format="PNG", optimize=True)
        saidas.append((buf.getvalue(), img.size[0], img.size[1]))
    return saidas


# ─── API pública ────────────────────────────────────────────────────────────


def gerar_capturas(
    xlsx_bytes: bytes,
    *,
    dpi: int = 150,
    nome_aba: str = "CÁLCULO",
    incluir_png_pagina_inteira: bool = True,
) -> CapturasGeradas:
    """Gera PDF + PNGs a partir de um XLSX em memória.

    Args:
        xlsx_bytes: bytes do XLSX (geralmente saída de `planilha.gerar_xlsx`).
        dpi: resolução das PNGs. 150 dá legibilidade sólida com peso ~300KB/página.
        nome_aba: nome da aba visual (usado pra compor nome dos PNGs).
        incluir_png_pagina_inteira: se False, só gera o PDF (economiza rasterização).

    Returns:
        `CapturasGeradas` com `pdf_bytes` e `pngs` (uma entrada por página de impressão).

    Raises:
        LibreOfficeIndisponivel: soffice não está no PATH.
        CapturasError: outra falha durante a conversão.
    """
    pdf_bytes = _xlsx_para_pdf(xlsx_bytes)
    if not incluir_png_pagina_inteira:
        logger.info(
            "Capturas: apenas PDF (%d KB), PNG de página inteira omitido.",
            len(pdf_bytes) // 1024,
        )
        return CapturasGeradas(pdf_bytes=pdf_bytes, pngs=[])

    paginas_raw = _pdf_para_pngs(pdf_bytes, dpi=dpi)

    pngs: list[CapturaPng] = []
    total = len(paginas_raw)
    for i, (png_bytes, w, h) in enumerate(paginas_raw, start=1):
        if total == 1:
            sufixo = ""
        else:
            sufixo = f" p{i}de{total}"
        nome = f"13 Print {nome_aba}{sufixo}.png"
        pngs.append(
            CapturaPng(
                nome_arquivo=nome,
                bytes_=png_bytes,
                pagina=i,
                largura_px=w,
                altura_px=h,
            )
        )

    logger.info(
        "Capturas geradas: PDF (%d KB) + %d PNG(s) total %d KB",
        len(pdf_bytes) // 1024,
        len(pngs),
        sum(len(p.bytes_) for p in pngs) // 1024,
    )
    return CapturasGeradas(pdf_bytes=pdf_bytes, pngs=pngs)


def nome_pdf(quitado: bool = False) -> str:
    """Nome do PDF unificado salvo na pasta da cliente.

    v0.9.0: nome sem numeração ("salvar como Calculo"); variante quitado mantém o sufixo.
    """
    return "Calculo quitado.pdf" if quitado else "Calculo.pdf"


def recortar_bloco_pmt_pdf(
    pdf_bytes: bytes,
    nome_arquivo: str,
    *,
    dpi: int = 200,
) -> CapturaPng | None:
    """Recorta o bloco INTEIRO "PARCELA COM TAXA MÉDIA E EXPURGO DE ABUSIVIDADES"
    do PDF do Cálculo já renderizado — fiel à planilha (sem re-render distorcido).

    v0.9.3: imag.02 = prova de como o cálculo foi feito (título + valores + a
    derivação da fórmula PMT). Recorta direto do PDF (sugestão da Rose: "print do
    PDF sai certinho"), evitando o `##`/gaps do re-render regional.

    Âncoras (presentes em todas as abas e nos fluxos ativo/quitado):
    - título: palavra "EXPURGO" (caixa-alta) → topo do bloco.
    - left/right: a BORDA da caixa que contém o título (rect do pdfplumber).
    - fim: topo do cabeçalho "PARCELAS RECALCULADAS" logo abaixo do bloco.

    Returns CapturaPng, ou None se as âncoras não baterem (fallback — nunca derruba o run).
    """
    import io

    try:
        import pdfplumber
    except ImportError as e:  # pragma: no cover
        raise CapturasError(f"pdfplumber não instalado: {e}") from e

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            titulos = [w for w in words if w["text"] == "EXPURGO"]
            if not titulos:
                continue
            titulo = min(titulos, key=lambda w: w["top"])
            # Caixa (rect) que contém o título → bordas left/right do bloco.
            caixas = [
                r for r in page.rects
                if r["top"] <= titulo["top"] + 4
                and r["bottom"] >= titulo["top"] - 14
                and r["x0"] < titulo["x0"] < r["x1"]
                and (r["x1"] - r["x0"]) > 150
            ]
            if not caixas:
                continue
            caixa = min(caixas, key=lambda r: abs(r["top"] - titulo["top"]))
            # Fim do bloco = topo do cabeçalho "PARCELAS RECALCULADAS".
            fins = [w for w in words if w["text"].upper() == "RECALCULADAS" and w["top"] > titulo["top"]]
            if not fins:
                continue
            bbox = (
                caixa["x0"] - 3,
                titulo["top"] - 8,
                caixa["x1"] + 3,
                min(w["top"] for w in fins) - 3,
            )
            img = page.within_bbox(bbox).to_image(resolution=dpi)
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            png = buf.getvalue()
            w_px, h_px = img.original.size
            logger.info("imag.02 (bloco PMT) recortada do PDF: %dx%d, %d KB", w_px, h_px, len(png) // 1024)
            return CapturaPng(nome_arquivo=nome_arquivo, bytes_=png, pagina=1, largura_px=w_px, altura_px=h_px)
    return None


# ─── Capturas regionais (XLSX por range) ────────────────────────────────────


def _trim_whitespace(png_bytes: bytes, *, padding_px: int = 12) -> tuple[bytes, int, int]:
    """Remove margem branca ao redor do conteúdo no PNG.

    Usa `Image.getbbox()` invertendo branco→preto pra detectar o bounding box
    do conteúdo. Adiciona `padding_px` ao redor pra não cortar bordas finas.

    Returns:
        (png_bytes_trimmed, largura, altura)
    """
    import io
    from PIL import Image, ImageChops

    img = Image.open(io.BytesIO(png_bytes)).convert("RGB")
    # Cria fundo branco da mesma dimensão e calcula diferença → não-branco
    bg = Image.new("RGB", img.size, (255, 255, 255))
    diff = ImageChops.difference(img, bg)
    bbox = diff.getbbox()
    if bbox is None:
        # Imagem totalmente branca — devolve original
        return png_bytes, img.size[0], img.size[1]

    left, top, right, bottom = bbox
    left = max(0, left - padding_px)
    top = max(0, top - padding_px)
    right = min(img.size[0], right + padding_px)
    bottom = min(img.size[1], bottom + padding_px)
    cropped = img.crop((left, top, right, bottom))

    buf = io.BytesIO()
    cropped.save(buf, format="PNG", optimize=True)
    return buf.getvalue(), cropped.size[0], cropped.size[1]


def gerar_capturas_regionais(
    xlsx_bytes: bytes,
    regioes: dict[str, str],
    *,
    nome_aba: str = "CÁLCULO",
    dpi: int = 150,
    trim: bool = True,
    nomes_arquivo: dict[str, str] | None = None,
) -> list[CapturaPng]:
    """Gera 1 PNG por região, alterando print_area temporariamente.

    Para cada (label, range_xlsx) em `regioes`:
    1. Carrega XLSX em memória
    2. Define `ws.print_area = range_xlsx`, limpa header/footer
    3. Salva XLSX modificado
    4. Converte com soffice → PDF (1 página)
    5. Rasteriza pra PNG
    6. Aplica trim de whitespace (se `trim=True`)
    7. Nome final: "13 Print {label}.png"

    Args:
        xlsx_bytes: bytes do XLSX original (saída de `planilha.gerar_xlsx`).
        regioes: dict {label_curto: range_xlsx (ex: "C6:Y18")}.
        nome_aba: aba visual de onde extrair as regiões.
        dpi: resolução das PNGs.
        trim: corta whitespace ao redor após render.

    Returns:
        Lista de `CapturaPng`, uma por região.
    """
    from io import BytesIO
    from openpyxl import load_workbook

    from openpyxl.utils import column_index_from_string, range_boundaries

    saidas: list[CapturaPng] = []
    for label, range_xlsx in regioes.items():
        # 1. Carrega + ajusta print_area
        wb = load_workbook(BytesIO(xlsx_bytes))
        if nome_aba not in wb.sheetnames:
            raise CapturasError(
                f"Aba {nome_aba!r} não encontrada no XLSX. Abas: {wb.sheetnames}"
            )
        ws = wb[nome_aba]
        # print_area é por-aba: o range basta (sem prefixo de aba). O prefixo
        # com nome contendo espaço — ex.: "PRICE 24X!AD25:AZ73" — quebra no
        # openpyxl ≥3.1.5 por falta de aspas. v0.9.0.
        ws.print_area = range_xlsx
        # Orientação adaptativa: se a região é mais ALTA que LARGA, usa portrait;
        # senão landscape. Evita que blocos quadrados/altos sejam comprimidos demais.
        col_min, row_min, col_max, row_max = range_boundaries(range_xlsx)
        n_cols = col_max - col_min + 1
        n_rows = row_max - row_min + 1
        # heurística: cells visuais ~ width 8 chars (~80 pt) × height 15 pt → ratio largura/altura ~5
        # se n_rows / n_cols > ~0.5, vira mais alto que largo em proporção visual
        if n_rows / max(1, n_cols) > 0.5:
            ws.page_setup.orientation = "portrait"
        else:
            ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 9  # A4

        # Rows compactas (~3pt) na região PMT: distribuir altura dentro do espaço
        # útil A4 portrait (~720pt) em vez de fixar 12/18pt para todas as linhas.
        AREA_UTIL_PT = 720
        rows_curtas_idx = [
            r
            for r in range(row_min, row_max + 1)
            if 0 < (ws.row_dimensions[r].height or 0) < 6
        ]
        rows_normais_total = 0.0
        for r in range(row_min, row_max + 1):
            if r in rows_curtas_idx:
                continue
            rh = ws.row_dimensions[r].height
            rows_normais_total += float(rh) if rh else 15.0
        espaco_para_curtas = AREA_UTIL_PT - rows_normais_total
        n_curtas = len(rows_curtas_idx)
        if n_curtas > 0 and espaco_para_curtas > 0:
            altura_curta = max(8, min(18, int(espaco_para_curtas // n_curtas)))
        else:
            altura_curta = 18 if n_curtas >= 5 else 12
        for r in range(row_min, row_max + 1):
            if r in rows_curtas_idx:
                ws.row_dimensions[r].height = altura_curta
            else:
                h = ws.row_dimensions[r].height
                if h is None or h < 12:
                    ws.row_dimensions[r].height = max(h or 0, 12)
        # Cabe em 1 página
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        # Limpa header/footer pra não aparecer "Cálculo de Ação Crefaz" / "Página 1 de 1"
        for hf_attr in ("oddHeader", "oddFooter", "evenHeader", "evenFooter"):
            hf = getattr(ws, hf_attr, None)
            if hf is not None:
                for align in ("left", "center", "right"):
                    section = getattr(hf, align, None)
                    if section is not None:
                        section.text = ""
        # Margens mínimas pra reduzir whitespace
        ws.page_margins.left = 0.1
        ws.page_margins.right = 0.1
        ws.page_margins.top = 0.1
        ws.page_margins.bottom = 0.1
        ws.page_margins.header = 0.0
        ws.page_margins.footer = 0.0
        wb.calculation.fullCalcOnLoad = True

        # 2. Salva e converte
        out = BytesIO()
        wb.save(out)
        xlsx_regional = out.getvalue()

        try:
            pdf_bytes = _xlsx_para_pdf(xlsx_regional)
            paginas_raw = _pdf_para_pngs(pdf_bytes, dpi=dpi)
        except CapturasError as e:
            logger.warning("Falha ao gerar região %s (%s): %s", label, range_xlsx, e)
            continue

        if not paginas_raw:
            logger.warning("Região %s gerou 0 páginas — ignorada.", label)
            continue

        # 3. Primeira página + trim
        png_bytes, w, h = paginas_raw[0]
        if trim:
            png_bytes, w, h = _trim_whitespace(png_bytes)
        nome = (nomes_arquivo or {}).get(label, f"13 Print {label}.png")
        saidas.append(
            CapturaPng(
                nome_arquivo=nome,
                bytes_=png_bytes,
                pagina=1,
                largura_px=w,
                altura_px=h,
            )
        )
        logger.info(
            "Região %s capturada: %s (%dx%d, %d KB)",
            label, nome, w, h, len(png_bytes) // 1024,
        )

    return saidas


# ─── Capturas de páginas de PDFs ────────────────────────────────────────────


def capturar_pagina_pdf(
    pdf_bytes: bytes,
    marcador_regex: str,
    nome_arquivo: str,
    *,
    dpi: int = 150,
    marcador_fim_regex: str | None = None,
    altura_fallback_ratio: float = 0.50,
) -> CapturaPng | None:
    """Captura como PNG a primeira página do PDF que contém o marcador.

    Args:
        pdf_bytes: bytes do PDF (contrato Crefaz ou Séries BACEN).
        marcador_regex: regex (re.IGNORECASE) buscado no texto extraído.
        nome_arquivo: nome do PNG resultante (ex: "13 Print 07 Item II do Contrato.png").
        dpi: resolução do PNG.

    Returns:
        `CapturaPng` da primeira página que casa, ou None se não encontrar.
    """
    import io
    import math
    import re
    try:
        import pypdfium2 as pdfium
    except ImportError as e:
        raise CapturasError(f"pypdfium2 não instalado: {e}") from e
    try:
        import pdfplumber
    except ImportError as e:
        raise CapturasError(f"pdfplumber não instalado: {e}") from e

    pattern = re.compile(marcador_regex, re.IGNORECASE)

    # Procura página por página + localização vertical dos marcadores
    # pdfplumber: origem no canto superior esquerdo, eixo y para baixo.
    # pypdfium2 crop: tupla (left, bottom, right, top) = quantidade recortada de cada borda,
    # em unidades do canvas PDF — NÃO é caixa (x0,y0,x1,y1). Ver PdfPage.render().
    pagina_alvo: int | None = None  # 0-indexed
    h_plumb: float | None = None
    y_inicio_plumb: float | None = None
    y_fim_plumb: float | None = None
    fim_pattern = re.compile(marcador_fim_regex, re.IGNORECASE) if marcador_fim_regex else None

    def _words_agrupados_em_linhas(words: list[dict]) -> list[tuple[float, str, list[float], list[float]]]:
        """Agrupa extract_words por faixa de top (~2pt); cada tupla = (bucket, texto_linha, tops, bottoms)."""
        buckets: dict[float, list[dict]] = {}
        for w in words:
            top = float(w.get("top", 0))
            b = round(top / 2.0) * 2.0
            buckets.setdefault(b, []).append(w)
        saida: list[tuple[float, str, list[float], list[float]]] = []
        for b in sorted(buckets.keys()):
            chunk = sorted(buckets[b], key=lambda x: float(x.get("x0", 0)))
            line_text = " ".join(x.get("text", "") for x in chunk).strip()
            tops = [float(x["top"]) for x in chunk]
            bottoms = [float(x["bottom"]) for x in chunk]
            saida.append((b, line_text, tops, bottoms))
        return saida

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for i, page in enumerate(pdf.pages):
            texto = page.extract_text() or ""
            if not pattern.search(texto):
                continue
            pagina_alvo = i
            h_plumb = float(page.height)
            try:
                words = page.extract_words() or []
            except Exception:
                words = []
            line_rows = _words_agrupados_em_linhas(words)
            for _bk, line_text, tops, _bottoms in line_rows:
                if pattern.search(line_text):
                    if tops:
                        y_inicio_plumb = max(0.0, min(tops) - 10.0)
                    break
            if y_inicio_plumb is None:
                y_inicio_plumb = 0.0
            if fim_pattern:
                for _bk, line_text, tops, bottoms in line_rows:
                    if not fim_pattern.search(line_text):
                        continue
                    if y_inicio_plumb is not None and tops and max(tops) <= y_inicio_plumb:
                        continue
                    if tops:
                        # v0.7.0: corta ACIMA do marcador de fim (ex.: "III. CUSTO
                        # EFETIVO TOTAL") — captura só a seção pedida (II), sem vazar
                        # o cabeçalho/tabela da seção seguinte. Antes usava
                        # max(bottoms)+12 e incluía a linha do III.
                        y_fim_plumb = max(y_inicio_plumb + 10.0, min(tops) - 4.0)
                        break
            break

    if pagina_alvo is None:
        logger.warning(
            "Marcador %r não encontrado no PDF — captura ignorada (%s).",
            marcador_regex, nome_arquivo,
        )
        return None

    # Renderiza a página alvo (apenas bloco, não página inteira)
    pdf = pdfium.PdfDocument(pdf_bytes)
    page = pdf[pagina_alvo]
    page_h = float(page.get_height())

    if y_inicio_plumb is None:
        y_inicio_plumb = 0.0
    if y_fim_plumb is None or y_fim_plumb <= y_inicio_plumb:
        y_fim_plumb = min(h_plumb or page_h, y_inicio_plumb + (h_plumb or page_h) * altura_fallback_ratio)

    y_inicio_plumb = max(0.0, min(y_inicio_plumb, h_plumb or page_h))
    y_fim_plumb = max(y_inicio_plumb + 10.0, min(y_fim_plumb, h_plumb or page_h))

    # Mapeia faixa vertical do pdfplumber para altura reportada pelo PDFium (podem diferir em pts).
    def _y_plumb_para_pdfium(y: float) -> float:
        hp = h_plumb if h_plumb > 0 else page_h
        if hp <= 0:
            return 0.0
        return (y / hp) * page_h

    y0 = _y_plumb_para_pdfium(y_inicio_plumb)
    y1 = _y_plumb_para_pdfium(y_fim_plumb)
    # Recortes: topo da página e base (unidades PDF, mesmas de get_width/get_height).
    top_cut = max(0.0, min(y0, page_h))
    bottom_cut = max(0.0, min(page_h - y1, page_h))
    # Replica o ceil interno do pypdfium2: soma dos cortes em px deve deixar altura ≥ 1.
    scale = dpi / 72.0
    src_height_px = math.ceil(page_h * scale)
    if src_height_px - math.ceil(bottom_cut * scale) - math.ceil(top_cut * scale) < 1:
        top_cut, bottom_cut = 0.0, 0.0

    img = page.render(
        scale=scale,
        crop=(0.0, bottom_cut, 0.0, top_cut),
    ).to_pil()
    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    png_bytes = buf.getvalue()

    logger.info(
        "PDF capturado: %s (página %d, %dx%d, %d KB)",
        nome_arquivo, pagina_alvo + 1, img.size[0], img.size[1], len(png_bytes) // 1024,
    )
    return CapturaPng(
        nome_arquivo=nome_arquivo,
        bytes_=png_bytes,
        pagina=pagina_alvo + 1,
        largura_px=img.size[0],
        altura_px=img.size[1],
    )
