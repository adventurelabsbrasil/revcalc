"""Preenche o template Calculo.xlsx com os dados do cálculo."""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from .calculadora import decidir_aba
from .config import (
    CELULAS_DADOS,
    NOME_ABA_CALCULO,
    NOME_ABA_DADOS,
    TEMPLATE_PATH,
    TEMPLATE_QUITADO_PATH,
    celulas_para_aba,
)
from .parser_contrato import DadosContrato

logger = logging.getLogger(__name__)

# Tabelas lado a lado na linha 130+ — mesmas linhas físicas para os 3 blocos.
_ROW_PARCELAS_INICIO = 132
_ROW_PARCELAS_FIM = 155
_REF_PARCELA_IDX = 131  # ROW()-131 = índice da parcela

# Range físico (início, fim) das 3 tabelas de parcelas por aba — usado para
# ocultar as linhas excedentes ao prazo (ver _ocultar_linhas_excedentes).
# ATIVO: aba CÁLCULO (cap. 24). QUITADO: abas PRICE, capacidade crescente +
# offset de layout (+1 linha nas longas). Parcela N ocupa a linha `início+N-1`.
_RANGE_PARCELAS_POR_ABA = {
    NOME_ABA_CALCULO: (132, 155),
    "PRICE 24X": (132, 155),
    "PRICE 36x": (133, 168),
    "PRICE 48x": (133, 180),
    "PRICE 60x": (133, 192),
}


@dataclass
class DadosPlanilha:
    """Dados que vão escritos no XLSX final."""

    contrato: DadosContrato
    taxa_bacen: float  # decimal, ex: 0.0647 para 6,47%
    parcelas_pagas: int
    data_calculo: date  # data_atual usada no cálculo


def _titulo_operacao(contrato: DadosContrato) -> str:
    return (
        f"CÁLCULOS DA OPERAÇÃO Nº {contrato.numero_cedula} - "
        f"CLIENTE: {contrato.nome_emitente.upper()} x BANCO CREFAZ"
    )


def _remover_abas_nao_usadas(
    wb: Workbook,
    aba_visual: str,
    *,
    manter: tuple[str, ...] = (),
) -> None:
    """Remove abas que não são a visual-alvo nem as extras (ex.: NOME_ABA_DADOS)."""
    for nome in list(wb.sheetnames):
        if nome == aba_visual or nome in manter:
            continue
        del wb[nome]


def _forcar_landscape(ws, aba: str) -> None:
    """Defesa em profundidade: força paisagem caso template seja modificado e perca a config.

    Em v0.6.0 o template `Calculo.xlsx` tem aba CÁLCULO única já em landscape A4
    com fitToWidth=1 e fitToHeight=1. Mantém função como backstop de runtime.
    """
    if ws.page_setup.orientation != "landscape":
        logger.warning(
            "Aba %s está em %s no template — forçando landscape.",
            aba,
            ws.page_setup.orientation,
        )
        ws.page_setup.orientation = "landscape"


def _set_date_cell(ws, coord: str, d: date) -> None:
    """Escreve data como datetime e força formato BR no XLSX."""
    cell = ws[coord]
    cell.value = datetime.combine(d, datetime.min.time())
    cell.number_format = "dd/mm/yyyy"


def _preencher_aba(ws, aba: str, dados: DadosPlanilha) -> None:
    celulas = celulas_para_aba(aba)
    contrato = dados.contrato

    # Título
    ws[celulas["titulo"]] = _titulo_operacao(contrato)

    # Datas (dd/mm/yyyy defensivo — não confiar só no template)
    _set_date_cell(ws, celulas["data_pactuacao"], contrato.data_emissao)
    _set_date_cell(ws, celulas["primeiro_vencimento"], contrato.primeiro_vencimento)

    # Valores monetários (R$)
    ws[celulas["valor_principal"]] = contrato.valor_nominal
    ws[celulas["tac"]] = contrato.tarifas
    ws[celulas["iof"]] = contrato.tributos_iof
    ws[celulas["qtd_parcelas"]] = contrato.prazo
    ws[celulas["valor_prestacao"]] = contrato.valor_prestacao

    # Taxas em decimal (formato % do template multiplica por 100 ao exibir)
    ws[celulas["taxa_pactuada"]] = contrato.taxa_mensal
    ws[celulas["taxa_bacen"]] = dados.taxa_bacen

    # Parcelas pagas (inteiro)
    ws[celulas["parcelas_pagas"]] = dados.parcelas_pagas


def _corrigir_tabelas_parcelas_e_colunas(ws, aba: str) -> None:
    """Alinha fórmulas ao template ativo: AD dinâmico; BE/BF só até parcelas pagas; colunas C–L visíveis.

    - PARCELAS RECALCULADAS (AD): índice como coluna C (sem números fixos).
    - VALORES PAGOS (BE/BF): só mostra linhas até `$BL$8` parcelas pagas.
    - Remove `hidden` em colunas C–L (defeito de impressão no template).
    - Oculta linha 5 (“Data do 1º vencimento”) — só datas pactuação/cálculo no cabeçalho.
    - Expoentes na linha visual da fórmula PMT: referência a `$I$15` (evita ## por coluna estreita).
    """
    if aba != NOME_ABA_CALCULO or ws.title != NOME_ABA_CALCULO:
        return
    idx = f"ROW()-{_REF_PARCELA_IDX}"
    for r in range(_ROW_PARCELAS_INICIO, _ROW_PARCELAS_FIM + 1):
        ws[f"AD{r}"] = f'=IF({idx}>$I$15,"",{idx})'
        ws[f"BE{r}"] = f'=IF({idx}>$I$15,"",IF({idx}<=$BL$8,{idx},""))'
        ws[f"BF{r}"] = f'=IF(BE{r}="","",AE{r})'
    for col in range(3, 13):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].hidden = False

    ws.row_dimensions[5].hidden = True

    # Expoentes na linha visual PMT: referência a $I$15 (evita ## por coluna estreita).
    ws["U41"] = "=$I$15"
    ws["AV41"] = "=$I$15"
    for letter in ("U", "V", "AV", "AW"):
        cur = ws.column_dimensions[letter].width
        ws.column_dimensions[letter].width = max(float(cur or 8), 10.0)


def _ocultar_linhas_excedentes(ws, aba: str, prazo: int) -> None:
    """Oculta as linhas das tabelas de parcelas além de `prazo` (ATIVO e QUITADO).

    As 3 tabelas (PARCELAS CONFORME O CONTRATO / RECALCULADAS / VALORES PAGOS)
    dividem um range físico fixo por aba; as fórmulas já esvaziam o excedente, mas
    as linhas em branco continuavam visíveis com bordas. Fecha a tabela em
    exatamente N linhas = nº de parcelas do contrato (feedback Rose/Bruna v0.9.7).
    Parcela N ocupa a linha `início + N - 1`; oculta de `início + prazo` até o fim.
    """
    faixa = _RANGE_PARCELAS_POR_ABA.get(aba)
    if not faixa:
        return
    inicio, fim = faixa
    primeira_oculta = inicio + prazo
    for r in range(inicio, fim + 1):
        ws.row_dimensions[r].hidden = r >= primeira_oculta


def _preencher_aba_dados(ws, dados: DadosPlanilha) -> None:
    """Escreve apenas na aba DADOS — a aba PRICE deve usar fórmulas =DADOS!..."""
    c = CELULAS_DADOS
    contrato = dados.contrato
    ws[c["titulo"]] = _titulo_operacao(contrato)
    _set_date_cell(ws, c["data_emissao"], contrato.data_emissao)
    _set_date_cell(ws, c["primeiro_vencimento"], contrato.primeiro_vencimento)
    _set_date_cell(ws, c["ultimo_vencimento"], contrato.ultimo_vencimento)
    ws[c["valor_principal"]] = contrato.valor_nominal
    ws[c["tac"]] = contrato.tarifas
    ws[c["iof"]] = contrato.tributos_iof
    ws[c["qtd_parcelas"]] = contrato.prazo
    ws[c["valor_prestacao"]] = contrato.valor_prestacao
    ws[c["taxa_pactuada"]] = contrato.taxa_mensal
    ws[c["taxa_bacen"]] = dados.taxa_bacen
    ws[c["parcelas_pagas"]] = dados.parcelas_pagas
    _set_date_cell(ws, c["data_calculo"], dados.data_calculo)


def gerar_xlsx(
    dados: DadosPlanilha,
    template_path: Path | None = None,
    *,
    quitado: bool = False,
) -> bytes:
    """Gera o XLSX preenchido em memória, retornando os bytes.

    - Carrega o template (ativo `Calculo.xlsx` ou quitado `Calculo.quitado.xlsx`)
    - Decide a aba apropriada via `decidir_aba(prazo, quitado)`
    - Remove abas não usadas (mantém `DADOS` se existir — modelo dados + visual)
    - Preenche: modo legacy (direto na aba) ou modo DADOS (só aba DADOS; visual com fórmulas)
    - Força landscape na aba visual se necessário
    - Retorna bytes prontos para upload

    No fluxo QUITADO o template não tem aba DADOS → cai no preenchimento legacy
    (`_preencher_aba`), escrevendo direto nas células da aba PRICE selecionada.
    """
    if template_path is None:
        template_path = TEMPLATE_QUITADO_PATH if quitado else TEMPLATE_PATH
    wb = load_workbook(template_path)

    aba = decidir_aba(dados.contrato.prazo, quitado)
    if aba not in wb.sheetnames:
        raise ValueError(
            f"Aba '{aba}' não encontrada no template. "
            f"Abas disponíveis: {wb.sheetnames}"
        )

    usa_dados = NOME_ABA_DADOS in wb.sheetnames
    manter = (NOME_ABA_DADOS,) if usa_dados else ()
    _remover_abas_nao_usadas(wb, aba, manter=manter)
    ws = wb[aba]

    if usa_dados:
        _preencher_aba_dados(wb[NOME_ABA_DADOS], dados)
        logger.info(
            "Template com aba %s — valores só na fonte; PRICE deve referenciar fórmulas.",
            NOME_ABA_DADOS,
        )
    else:
        _preencher_aba(ws, aba, dados)

    _forcar_landscape(ws, aba)
    _corrigir_tabelas_parcelas_e_colunas(ws, aba)
    _ocultar_linhas_excedentes(ws, aba, dados.contrato.prazo)

    # Força recálculo na abertura — sem isso, openpyxl salva o XLSX sem
    # calcChain.xml e a primeira renderização (LibreOffice headless ou
    # Sheets/Excel cliente) pode mostrar células de fórmula vazias até o
    # usuário tocar/recalcular manualmente.
    wb.calculation.fullCalcOnLoad = True

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
