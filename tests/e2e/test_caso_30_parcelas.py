"""Smoke B7 — 30 parcelas em PRICE 36x (vencimentos coluna D via fórmula)."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from calculadora_crefaz.parser_contrato import DadosContrato
from calculadora_crefaz.planilha import DadosPlanilha, gerar_xlsx


def _contrato_30_mock() -> DadosContrato:
    return DadosContrato(
        numero_cedula="9999999",
        nome_emitente="Cliente Hipotético Trinta Parcelas",
        data_emissao=date(2026, 4, 1),
        primeiro_vencimento=date(2026, 5, 1),
        ultimo_vencimento=date(2028, 10, 1),
        valor_nominal=12000.00,
        valor_emprestimo=12500.00,
        valor_total_contratado=25000.00,
        valor_prestacao=500.00,
        tributos_iof=500.00,
        tarifas=0.00,
        prazo=30,
        taxa_mensal=0.05,
        taxa_anual=0.80,
    )


def test_30_parcelas_price_36x_datas_e_formulas():
    dados = DadosPlanilha(
        contrato=_contrato_30_mock(),
        taxa_bacen=0.02,
        parcelas_pagas=0,
        data_calculo=date(2026, 4, 28),
    )
    blob = gerar_xlsx(dados)
    wb = load_workbook(BytesIO(blob))

    assert wb.sheetnames == ["PRICE 36x"]
    ws = wb.active

    assert ws["D6"].value.date() == date(2026, 5, 1)
    assert ws["D6"].number_format == "dd/mm/yyyy"
    assert ws["I16"].value == 30
    assert ws["I17"].value == 500.00

    assert str(ws["D133"].value) == '=IF(C133<=$I$16, $D$6, "")'
    assert "EDATE($D$6" in str(ws["D162"].value)
    assert "EDATE($D$6" in str(ws["D163"].value)

    assert str(ws["N133"].value) == '=IF(C133<=$I$16, $I$17, "")'
    assert "EDATE" not in str(ws["N162"].value)

    assert ws.page_setup.orientation == "landscape"
    assert ws.page_setup.fitToWidth in (1, True)
