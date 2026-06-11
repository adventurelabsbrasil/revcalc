"""Smoke e2e — gerar_xlsx num caso ≤24 parcelas.

Valida a aba CÁLCULO única (modelo DADOS + visual), o wiring por fórmula p/ a
aba DADOS, e as fórmulas dinâmicas da tabela de parcelas (AD/BE/BF) que o engine
injeta em `_corrigir_tabelas_parcelas_e_colunas`. v0.6.0 aposentou as abas
multi-prazo (PRICE 36x/48x/60x) — Crefaz só opera até 24 parcelas.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from calculadora_crefaz.parser_contrato import DadosContrato
from calculadora_crefaz.planilha import DadosPlanilha, gerar_xlsx


def _contrato_18_mock() -> DadosContrato:
    return DadosContrato(
        numero_cedula="9999999",
        nome_emitente="Cliente Hipotético Dezoito Parcelas",
        data_emissao=date(2026, 1, 1),
        primeiro_vencimento=date(2026, 2, 1),
        ultimo_vencimento=date(2027, 7, 1),
        valor_nominal=8000.00,
        valor_emprestimo=8500.00,
        valor_total_contratado=15000.00,
        valor_prestacao=600.00,
        tributos_iof=500.00,
        tarifas=0.00,
        prazo=18,
        taxa_mensal=0.05,
        taxa_anual=0.80,
    )


def test_caso_18_parcelas_aba_calculo_e_formulas_dinamicas():
    dados = DadosPlanilha(
        contrato=_contrato_18_mock(),
        taxa_bacen=0.02,
        parcelas_pagas=4,
        data_calculo=date(2026, 4, 28),
    )
    wb = load_workbook(BytesIO(gerar_xlsx(dados)))

    assert wb.sheetnames == ["DADOS", "CÁLCULO"]
    ws = wb["CÁLCULO"]

    # Inputs referenciam a aba DADOS.
    assert ws["I15"].value == "=DADOS!$B$9"   # qtd parcelas
    assert ws["BL8"].value == "=DADOS!$B$13"  # parcelas pagas

    # Tabela de parcelas: fórmulas dinâmicas (índice = ROW()-131), idênticas nas
    # linhas 132–155, limitadas por $I$15 (qtd parcelas) e $BL$8 (parcelas pagas).
    assert ws["AD132"].value == '=IF(ROW()-131>$I$15,"",ROW()-131)'
    assert ws["AD155"].value == '=IF(ROW()-131>$I$15,"",ROW()-131)'
    assert ws["BE132"].value == '=IF(ROW()-131>$I$15,"",IF(ROW()-131<=$BL$8,ROW()-131,""))'
    assert ws["BF132"].value == '=IF(BE132="","",AE132)'

    # Expoente da fórmula visual PMT aponta p/ $I$15 (evita ## por coluna estreita).
    assert ws["U41"].value == "=$I$15"
    assert ws["AV41"].value == "=$I$15"

    # Linha 5 ("Data do 1º vencimento") oculta; impressão landscape A4.
    assert ws.row_dimensions[5].hidden is True
    assert ws.page_setup.orientation == "landscape"
    assert ws.page_setup.fitToWidth in (1, True)
