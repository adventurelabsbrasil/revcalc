"""Testes do preenchimento da planilha — usam o template real do projeto."""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from calculadora_crefaz.exceptions import PrazoForaDoTemplate
from calculadora_crefaz.parser_contrato import DadosContrato
from calculadora_crefaz.planilha import DadosPlanilha, gerar_xlsx


def _contrato_marli() -> DadosContrato:
    return DadosContrato(
        numero_cedula="4095068",
        nome_emitente="Marli Sueli Berger Dambrosio",
        data_emissao=date(2025, 12, 29),
        primeiro_vencimento=date(2026, 2, 2),
        ultimo_vencimento=date(2027, 7, 2),
        valor_nominal=3500.00,
        valor_emprestimo=3604.99,
        valor_total_contratado=10539.54,
        valor_prestacao=585.53,
        tributos_iof=104.99,
        tarifas=0.00,
        prazo=18,
        taxa_mensal=0.1449,
        taxa_anual=3.9373,
    )


def _contrato_adriano() -> DadosContrato:
    return DadosContrato(
        numero_cedula="3867296",
        nome_emitente="Adriano Luis Calistro Lourenco",
        data_emissao=date(2025, 9, 22),
        primeiro_vencimento=date(2025, 10, 27),
        ultimo_vencimento=date(2026, 9, 27),
        valor_nominal=1000.00,
        valor_emprestimo=1025.10,
        valor_total_contratado=2721.48,
        valor_prestacao=226.79,
        tributos_iof=25.10,
        tarifas=0.00,
        prazo=12,
        taxa_mensal=0.1877,
        taxa_anual=6.9289,
    )


def _carrega_xlsx(blob: bytes):
    return load_workbook(BytesIO(blob))


class TestGerarXlsx:
    # v0.6.0: o template Calculo.xlsx tem aba DADOS (valores) + aba CÁLCULO única
    # (fórmulas =DADOS!...). Os inputs vão na DADOS; a CÁLCULO referencia. Crefaz só
    # opera até 24 parcelas → abas multi-prazo (PRICE 36x/48x/60x) foram removidas.
    def test_marli_18_parcelas_dados_e_calculo(self):
        dados = DadosPlanilha(
            contrato=_contrato_marli(),
            taxa_bacen=0.0647,
            parcelas_pagas=3,
            data_calculo=date(2026, 4, 28),
        )
        wb = _carrega_xlsx(gerar_xlsx(dados))

        assert wb.sheetnames == ["DADOS", "CÁLCULO"]
        assert wb.active.title == "CÁLCULO"

        # Inputs vão na aba DADOS (ver CELULAS_DADOS).
        dd = wb["DADOS"]
        assert dd["B2"].value == (
            "CÁLCULOS DA OPERAÇÃO Nº 4095068 - "
            "CLIENTE: MARLI SUELI BERGER DAMBROSIO x BANCO CREFAZ"
        )
        assert dd["B3"].value.date() == date(2025, 12, 29)  # data_emissao
        assert dd["B4"].value.date() == date(2026, 2, 2)    # 1º vencimento
        assert dd["B4"].number_format == "dd/mm/yyyy"
        assert dd["B6"].value == 3500.00                    # valor principal
        assert dd["B7"].value == 0.00                       # tac/tarifas
        assert dd["B8"].value == 104.99                     # IOF
        assert dd["B9"].value == 18                         # qtd parcelas
        assert dd["B10"].value == 585.53                    # prestação
        assert dd["B11"].value == pytest.approx(0.1449)     # taxa pactuada
        assert dd["B12"].value == pytest.approx(0.0647)     # taxa BACEN
        assert dd["B13"].value == 3                         # parcelas pagas

        # A aba CÁLCULO referencia a DADOS por fórmula (não tem valores literais).
        cc = wb["CÁLCULO"]
        assert cc["C1"].value == "=DADOS!$B$2"
        assert cc["I7"].value == "=DADOS!$B$6"
        assert cc["I15"].value == "=DADOS!$B$9"
        assert cc["BL8"].value == "=DADOS!$B$13"
        assert cc.page_setup.orientation == "landscape"
        assert cc.page_setup.fitToWidth in (1, True)

    def test_adriano_12_parcelas_dados(self):
        dados = DadosPlanilha(
            contrato=_contrato_adriano(),
            taxa_bacen=0.0573,
            parcelas_pagas=6,
            data_calculo=date(2026, 4, 28),
        )
        wb = _carrega_xlsx(gerar_xlsx(dados))

        assert wb.sheetnames == ["DADOS", "CÁLCULO"]
        dd = wb["DADOS"]
        assert dd["B2"].value == (
            "CÁLCULOS DA OPERAÇÃO Nº 3867296 - "
            "CLIENTE: ADRIANO LUIS CALISTRO LOURENCO x BANCO CREFAZ"
        )
        assert dd["B4"].value.date() == date(2025, 10, 27)
        assert dd["B4"].number_format == "dd/mm/yyyy"
        assert dd["B6"].value == 1000.00
        assert dd["B8"].value == 25.10
        assert dd["B9"].value == 12
        assert dd["B10"].value == 226.79
        assert dd["B11"].value == pytest.approx(0.1877)
        assert dd["B13"].value == 6

    def test_prazo_acima_de_24_lanca(self):
        # Crefaz não opera > 24 parcelas; gerar_xlsx propaga PrazoForaDoTemplate.
        contrato = _contrato_marli()
        contrato.prazo = 30
        dados = DadosPlanilha(
            contrato=contrato,
            taxa_bacen=0.0647,
            parcelas_pagas=5,
            data_calculo=date(2026, 4, 28),
        )
        with pytest.raises(PrazoForaDoTemplate):
            gerar_xlsx(dados)

    def test_formulas_preservadas(self):
        """Garante que fórmulas críticas do template (aba CÁLCULO) não são sobrescritas."""
        dados = DadosPlanilha(
            contrato=_contrato_marli(),
            taxa_bacen=0.0647,
            parcelas_pagas=3,
            data_calculo=date(2026, 4, 28),
        )
        wb = _carrega_xlsx(gerar_xlsx(dados))
        ws = wb["CÁLCULO"]

        # Fórmulas-chave que NÃO devem ter sido tocadas pelo preenchimento.
        assert ws["I14"].value == "=SUM(I7:Y13)"  # total
        assert ws["I18"].value == "=I15*I16"  # valor final
        assert ws["AP14"].value == "=SUM(AP7:AZ13)"
        assert ws["AP17"].value == "=I15*AP16"
        assert ws["BL10"].value == "=BL8*BL9"


# ─── Fluxo QUITADO (v0.9.0) — template Calculo.quitado.xlsx ───────────────────


class TestGerarXlsxQuitado:
    def test_24x_inputs_e_aba_unica(self):
        """Quitado prazo ≤24 → aba PRICE 24X, preenchimento legacy direto nas células I."""
        dados = DadosPlanilha(
            contrato=_contrato_marli(),  # prazo 18
            taxa_bacen=0.0647,
            parcelas_pagas=18,  # todas pagas
            data_calculo=date(2026, 4, 28),
        )
        wb = _carrega_xlsx(gerar_xlsx(dados, quitado=True))

        # Só a aba selecionada sobra (sem DADOS — template quitado é legacy).
        assert wb.sheetnames == ["PRICE 24X"]
        ws = wb["PRICE 24X"]

        # Inputs escritos direto (CELULAS_CALCULO, offset 0).
        assert ws["I7"].value == 3500.00          # principal
        assert ws["I9"].value == 0.00             # TAC/tarifas
        assert ws["I13"].value == 104.99          # IOF
        assert ws["I15"].value == 18              # qtd parcelas
        assert ws["I16"].value == 585.53          # prestação
        assert ws["I17"].value == pytest.approx(0.1449)
        assert ws["BL8"].value == 18              # parcelas pagas
        assert ws["D5"].value.date() == date(2026, 2, 2)  # 1º vencimento (âncora EDATE)

        # Lista de parcelas guardada + datas via EDATE até a capacidade (linha 155).
        assert ws["C132"].value == '=IF(ROW()-131>$I$15,"",ROW()-131)'
        assert ws["D155"].value == '=IF(ROW()-131>$I$15,"",EDATE($D$5,ROW()-132))'

        # imag.02 resumo (v0.9.2): linha "Valor da parcela recalculada" injetada.
        assert "recalculada" in (ws["AD32"].value or "")
        assert ws["AI32"].value == "=AP16"

    def test_36x_offset_de_celulas(self):
        """Quitado prazo 25–36 → aba PRICE 36x com offset de linha (+1)."""
        contrato = _contrato_marli()
        contrato.prazo = 30
        dados = DadosPlanilha(
            contrato=contrato,
            taxa_bacen=0.0647,
            parcelas_pagas=30,
            data_calculo=date(2026, 4, 28),
        )
        wb = _carrega_xlsx(gerar_xlsx(dados, quitado=True))
        assert wb.sheetnames == ["PRICE 36x"]
        ws = wb["PRICE 36x"]
        # Offset +1: qtd em I16, prestação I17, parcelas pagas BL9, 1º venc D6.
        assert ws["I8"].value == 3500.00   # principal (I7→I8)
        assert ws["I16"].value == 30       # qtd parcelas (I15→I16)
        assert ws["BL9"].value == 30       # parcelas pagas (BL8→BL9)
        assert ws["D6"].value.date() == date(2026, 2, 2)  # 1º venc (D5→D6)

    def test_prazo_acima_de_60_lanca(self):
        contrato = _contrato_marli()
        contrato.prazo = 61
        dados = DadosPlanilha(
            contrato=contrato,
            taxa_bacen=0.0647,
            parcelas_pagas=61,
            data_calculo=date(2026, 4, 28),
        )
        with pytest.raises(PrazoForaDoTemplate):
            gerar_xlsx(dados, quitado=True)
