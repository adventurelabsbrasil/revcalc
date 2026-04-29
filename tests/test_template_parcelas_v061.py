"""Garantias mínimas de fórmulas nas três tabelas de parcelas (template Calculo.xlsx)."""

from __future__ import annotations

import re

from openpyxl import load_workbook

from calculadora_crefaz.config import TEMPLATE_PATH


def test_coluna_parcela_nao_hardcoded_ate_24() -> None:
    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb["CÁLCULO"]
    c132 = ws["C132"].value
    assert isinstance(c132, str) and "$I$15" in c132 and "ROW()" in c132.upper()
    assert not re.fullmatch(r"\d+", str(ws["C143"].value or "").strip())


def test_valores_pagos_so_ate_parcelas_pagas_na_col_bp() -> None:
    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb["CÁLCULO"]
    bp132 = ws["BP132"].value
    assert isinstance(bp132, str) and "$BL$8" in bp132


def test_ao_corta_apos_prazo() -> None:
    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb["CÁLCULO"]
    ao143 = ws["AO143"].value
    ao144 = ws["AO144"].value
    assert isinstance(ao143, str) and "$I$15" in ao143 and "AO142" in ao143
    assert isinstance(ao144, str) and "$I$15" in ao144
